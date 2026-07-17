from __future__ import annotations

import ast
import hashlib
import json
from email import policy
from email.parser import BytesParser
from pathlib import Path

import pytest
from icalendar import Calendar
from openpyxl import load_workbook
from pypdf import PdfReader

from continuity_ai.artifact_io import GroundTruthAccessError, open_production_artifact
from continuity_ai.aurora_fixture import ARTIFACTS, generate_project_aurora_fixture, manifest

ARTIFACT_ROOT = Path("fixtures/project_aurora/generated/artifacts")
TEST_ONLY_ROOT = Path("fixtures/project_aurora/generated/test_only")


def test_generates_all_required_artifacts(tmp_path: Path) -> None:
    generate_project_aurora_fixture(tmp_path)
    expected_paths = {artifact.relative_path for artifact in ARTIFACTS} | {
        str(TEST_ONLY_ROOT / "ground_truth.json")
    }
    assert expected_paths == {item["path"] for item in _all_generated(tmp_path)}
    for relative_path in expected_paths:
        assert (tmp_path / relative_path).is_file()


def test_production_artifact_input_contains_no_ground_truth(tmp_path: Path) -> None:
    generate_project_aurora_fixture(tmp_path)
    production_root = tmp_path / ARTIFACT_ROOT
    assert production_root.is_dir()
    assert list(production_root.rglob("ground_truth.json")) == []
    assert (tmp_path / TEST_ONLY_ROOT / "ground_truth.json").is_file()


def test_required_source_ids_exist_with_metadata(tmp_path: Path) -> None:
    generate_project_aurora_fixture(tmp_path)
    data = manifest(tmp_path)
    source_ids = {artifact["source_id"] for artifact in data["artifacts"]}
    assert source_ids == {
        "aurora-email-investor-approval-001",
        "aurora-calendar-production-001",
        "aurora-budget-v4-001",
        "aurora-callsheet-current-001",
        "aurora-crew-briefing-note-001",
    }
    for artifact in data["artifacts"]:
        assert artifact["evidence_id"].startswith("EV-AUR-")
        assert artifact["author"]
        assert artifact["timestamp"].endswith("Z")
        assert artifact["source_type"] in {"email", "calendar", "spreadsheet", "pdf", "markdown"}
        assert isinstance(artifact["timeline_position"], int)
        assert artifact["business_purpose"]
        assert (tmp_path / artifact["uri"]).is_file()
        assert len(artifact["sha256"]) == 64


def test_files_are_parseable_with_real_parsers(tmp_path: Path) -> None:
    generate_project_aurora_fixture(tmp_path)

    email_message = BytesParser(policy=policy.default).parsebytes(
        (tmp_path / ARTIFACT_ROOT / "email/investor_approval.eml").read_bytes()
    )
    assert email_message["Subject"] == "Approved: Project Aurora move to Northlight Studio"
    assert "Northlight Studio" in email_message.get_content()

    calendar = Calendar.from_ical((tmp_path / ARTIFACT_ROOT / "calendar/production_calendar.ics").read_bytes())
    events = [component for component in calendar.walk() if component.name == "VEVENT"]
    assert len(events) == 1
    assert str(events[0].get("LOCATION")) == "Harbor House"

    workbook = load_workbook(tmp_path / ARTIFACT_ROOT / "budget/budget_v4.xlsx", data_only=True)
    values = [row[0] for row in workbook["Budget v4"].iter_rows(values_only=True)]
    assert "Northlight Studio rental" in values

    reader = PdfReader(tmp_path / ARTIFACT_ROOT / "call_sheets/current_call_sheet.pdf")
    assert len(reader.pages) == 1
    text = reader.pages[0].extract_text()
    assert "Location: Harbor House" in text

    note = (tmp_path / ARTIFACT_ROOT / "notes/crew_briefing.md").read_text()
    assert "Briefing date: 2026-07-17" in note


def test_two_independent_generations_are_byte_identical(tmp_path: Path) -> None:
    first_root = tmp_path / "first"
    second_root = tmp_path / "second"
    generate_project_aurora_fixture(first_root)
    generate_project_aurora_fixture(second_root)
    assert _checksums(first_root) == _checksums(second_root)


def test_ground_truth_contains_expected_break_evidence_and_next_action(tmp_path: Path) -> None:
    generate_project_aurora_fixture(tmp_path)
    truth = json.loads((tmp_path / TEST_ONLY_ROOT / "ground_truth.json").read_text())
    assert truth["continuity_break"] == (
        "The approved location change is reflected in the budget but not in the production calendar or current call sheet."
    )
    assert truth["required_evidence"] == [
        "aurora-email-investor-approval-001",
        "aurora-budget-v4-001",
        "aurora-calendar-production-001",
        "aurora-callsheet-current-001",
    ]
    assert truth["expected_next_action"] == (
        "Update the production calendar and call sheet before tomorrow's crew briefing."
    )


def test_reasoning_modules_do_not_name_ground_truth_path() -> None:
    for path in Path("src/continuity_ai").glob("*reasoning*.py"):
        tree = ast.parse(path.read_text(), filename=str(path))
        string_constants = [node.value for node in ast.walk(tree) if isinstance(node, ast.Constant) and isinstance(node.value, str)]
        assert all("ground_truth.json" not in value for value in string_constants)


def test_runtime_guard_blocks_only_ground_truth_file(tmp_path: Path) -> None:
    artifact = tmp_path / "call_sheet.pdf"
    artifact.write_bytes(b"not a real pdf but allowed by the ground-truth guard")
    with open_production_artifact(artifact) as handle:
        assert handle.read() == artifact.read_bytes()

    blocked = tmp_path / "ground_truth.json"
    blocked.write_text("{}")
    with pytest.raises(GroundTruthAccessError):
        with open_production_artifact(blocked):
            pass


def _all_generated(root: Path) -> list[dict[str, str]]:
    return [
        {"path": str(path.relative_to(root)), "sha256": hashlib.sha256(path.read_bytes()).hexdigest()}
        for path in sorted((root / "fixtures/project_aurora/generated").rglob("*"))
        if path.is_file()
    ]


def _checksums(root: Path) -> dict[str, str]:
    return {item["path"]: item["sha256"] for item in _all_generated(root)}
