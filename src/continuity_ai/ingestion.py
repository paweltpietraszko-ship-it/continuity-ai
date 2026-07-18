"""Deterministic ingestion and normalization of Project Aurora production artifacts.

This module converts the production artifact directory into typed, normalized
evidence records. It performs no AI reasoning, contradiction detection,
summarization, or next-action generation, and it never reads test-only ground
truth. It must not depend on the fixture generator's fixed artifact
definitions, so that ingestion is verified purely against the on-disk
production contract (the evidence manifest), not against in-process constants.

The production evidence contract intentionally excludes interpretive fixture
fields (such as timeline position or business purpose): chronology is derived
solely from each artifact's own validated timestamp, never from a
manifest-provided ordering hint.
"""

from __future__ import annotations

import hashlib
import io
import json
import re
from datetime import datetime, timezone
from email import policy
from email.parser import BytesParser
from pathlib import Path, PurePosixPath, PureWindowsPath
from typing import Any

from icalendar import Calendar
from openpyxl import load_workbook
from pypdf import PdfReader

from continuity_ai.artifact_io import validate_production_artifact_root
from continuity_ai.models import EvidenceRecord

MANIFEST_FILENAME = "evidence_manifest.json"

_REQUIRED_ARTIFACT_FIELDS = (
    "source_id",
    "evidence_id",
    "author",
    "timestamp",
    "source_type",
    "title",
    "uri",
    "sha256",
)
_ALLOWED_ARTIFACT_FIELDS = set(_REQUIRED_ARTIFACT_FIELDS)
_ALLOWED_TOP_LEVEL_FIELDS = {"schema_version", "project", "artifacts"}

_SHA256_PATTERN = re.compile(r"^[0-9a-fA-F]{64}$")

_FORBIDDEN_DIRECTORY_NAME = "test_only"
_FORBIDDEN_FILENAME = "ground_truth.json"


class ArtifactIngestionError(RuntimeError):
    """Raised when Project Aurora artifact ingestion cannot proceed safely."""


def ingest_artifacts(artifact_root: Path) -> tuple[EvidenceRecord, ...]:
    """Read, validate, and normalize production artifacts into typed evidence records."""

    validate_production_artifact_root(artifact_root)

    manifest_path = artifact_root / MANIFEST_FILENAME
    payload = _load_manifest(manifest_path)
    entries = _validate_manifest_entries(payload)

    ordered: list[tuple[datetime, str, EvidenceRecord]] = []
    for entry in entries:
        resolved_path = _resolve_and_validate_uri(artifact_root, entry["uri"])
        raw_bytes = _read_verified_bytes(resolved_path, entry["sha256"])
        content = _parse_artifact(entry["source_type"], raw_bytes)
        record = EvidenceRecord(
            source_id=entry["source_id"],
            evidence_id=entry["evidence_id"],
            author=entry["author"],
            timestamp=entry["timestamp"],
            source_type=entry["source_type"],
            title=entry["title"],
            uri=entry["uri"],
            artifact_sha256=entry["sha256"],
            content=content,
        )
        ordered.append((entry["_parsed_timestamp"], entry["evidence_id"], record))

    ordered.sort(key=lambda item: (item[0], item[1]))
    return tuple(record for _, _, record in ordered)


def read_project_name(artifact_root: Path) -> str:
    """Return the manifest's project name, independently validated the same way
    ingest_artifacts validates the rest of the manifest, so callers can establish
    project identity without depending on ingest_artifacts's evidence-record shape."""
    validate_production_artifact_root(artifact_root)
    manifest_path = artifact_root / MANIFEST_FILENAME
    payload = _load_manifest(manifest_path)
    _validate_manifest_entries(payload)
    return payload["project"]


def _load_manifest(manifest_path: Path) -> Any:
    if not manifest_path.is_file():
        raise ArtifactIngestionError(f"Evidence manifest not found at {manifest_path}.")
    try:
        raw_bytes = manifest_path.read_bytes()
    except OSError as exc:
        raise ArtifactIngestionError(f"Evidence manifest at {manifest_path} could not be read.") from exc
    try:
        text = raw_bytes.decode("utf-8")
    except UnicodeDecodeError as exc:
        raise ArtifactIngestionError(f"Evidence manifest at {manifest_path} is not valid UTF-8.") from exc
    try:
        return json.loads(text)
    except json.JSONDecodeError as exc:
        raise ArtifactIngestionError(f"Evidence manifest at {manifest_path} is not valid JSON.") from exc


def _validate_manifest_entries(payload: Any) -> list[dict[str, Any]]:
    if not isinstance(payload, dict):
        raise ArtifactIngestionError("Evidence manifest must be a JSON object.")

    extra_top_level = set(payload.keys()) - _ALLOWED_TOP_LEVEL_FIELDS
    if extra_top_level:
        raise ArtifactIngestionError(
            f"Evidence manifest has unexpected top-level fields: {sorted(extra_top_level)}."
        )
    if payload.get("schema_version") != 1:
        raise ArtifactIngestionError("Evidence manifest schema_version is missing or unsupported.")
    if not isinstance(payload.get("project"), str) or not payload["project"]:
        raise ArtifactIngestionError("Evidence manifest project field is missing or invalid.")

    artifacts = payload.get("artifacts")
    if not isinstance(artifacts, list) or not artifacts:
        raise ArtifactIngestionError("Evidence manifest artifacts field is missing or empty.")

    validated: list[dict[str, Any]] = []
    seen_source_ids: set[str] = set()
    seen_evidence_ids: set[str] = set()
    seen_uris_lower: set[str] = set()

    for entry in artifacts:
        if not isinstance(entry, dict):
            raise ArtifactIngestionError("Evidence manifest artifact entries must be JSON objects.")

        extra_fields = set(entry.keys()) - _ALLOWED_ARTIFACT_FIELDS
        if extra_fields:
            raise ArtifactIngestionError(
                f"Evidence manifest artifact entry has unexpected fields: {sorted(extra_fields)}."
            )
        for field in _REQUIRED_ARTIFACT_FIELDS:
            if field not in entry:
                raise ArtifactIngestionError(
                    f"Evidence manifest artifact entry is missing required field '{field}'."
                )

        source_id = entry["source_id"]
        evidence_id = entry["evidence_id"]
        source_type = entry["source_type"]
        uri = entry["uri"]
        sha256 = entry["sha256"]

        if not isinstance(source_id, str) or not source_id:
            raise ArtifactIngestionError("Evidence manifest source_id must be a non-empty string.")
        if not isinstance(evidence_id, str) or not evidence_id:
            raise ArtifactIngestionError("Evidence manifest evidence_id must be a non-empty string.")
        if not isinstance(entry["author"], str) or not entry["author"]:
            raise ArtifactIngestionError("Evidence manifest author must be a non-empty string.")
        if not isinstance(source_type, str) or source_type not in _PARSERS:
            raise ArtifactIngestionError(f"Evidence manifest source_type '{source_type}' is not supported.")
        if not isinstance(entry["title"], str) or not entry["title"]:
            raise ArtifactIngestionError("Evidence manifest title must be a non-empty string.")
        if not isinstance(uri, str) or not uri:
            raise ArtifactIngestionError("Evidence manifest uri must be a non-empty string.")
        if not isinstance(sha256, str) or not _SHA256_PATTERN.match(sha256):
            raise ArtifactIngestionError("Evidence manifest sha256 must be a 64-character hexadecimal string.")

        canonical_timestamp, parsed_timestamp = _validate_timestamp(entry["timestamp"], evidence_id)

        if source_id in seen_source_ids:
            raise ArtifactIngestionError(f"Duplicate source_id '{source_id}' in evidence manifest.")
        if evidence_id in seen_evidence_ids:
            raise ArtifactIngestionError(f"Duplicate evidence_id '{evidence_id}' in evidence manifest.")
        uri_key = uri.lower()
        if uri_key in seen_uris_lower:
            raise ArtifactIngestionError(f"Duplicate uri '{uri}' in evidence manifest (case-insensitive).")
        seen_source_ids.add(source_id)
        seen_evidence_ids.add(evidence_id)
        seen_uris_lower.add(uri_key)

        _reject_unsafe_uri(uri)

        validated_entry = dict(entry)
        validated_entry["timestamp"] = canonical_timestamp
        validated_entry["_parsed_timestamp"] = parsed_timestamp
        validated.append(validated_entry)

    return validated


def _validate_timestamp(value: Any, evidence_id: str) -> tuple[str, datetime]:
    if not isinstance(value, str) or not value:
        raise ArtifactIngestionError(
            f"Evidence manifest timestamp for '{evidence_id}' must be a non-empty string."
        )
    try:
        parsed = datetime.fromisoformat(value)
    except ValueError as exc:
        raise ArtifactIngestionError(
            f"Evidence manifest timestamp '{value}' for '{evidence_id}' is not valid ISO 8601 / RFC 3339."
        ) from exc
    if parsed.tzinfo is None or parsed.utcoffset() is None:
        raise ArtifactIngestionError(
            f"Evidence manifest timestamp '{value}' for '{evidence_id}' must include explicit timezone information."
        )
    utc_parsed = parsed.astimezone(timezone.utc)
    canonical = utc_parsed.isoformat().replace("+00:00", "Z")
    return canonical, utc_parsed


def _reject_unsafe_uri(uri: str) -> None:
    if "\\" in uri:
        raise ArtifactIngestionError(f"Evidence manifest uri '{uri}' must use forward slashes only.")

    windows_uri = PureWindowsPath(uri)
    if windows_uri.drive or windows_uri.is_absolute():
        raise ArtifactIngestionError(f"Evidence manifest uri '{uri}' must not use a Windows drive or be absolute.")

    posix_uri = PurePosixPath(uri)
    if posix_uri.is_absolute():
        raise ArtifactIngestionError(f"Evidence manifest uri '{uri}' must not be absolute.")
    if any(part == ".." for part in posix_uri.parts):
        raise ArtifactIngestionError(f"Evidence manifest uri '{uri}' must not contain path traversal.")

    _reject_forbidden_ground_truth_reference(uri, posix_uri)


def _reject_forbidden_ground_truth_reference(uri: str, posix_uri: PurePosixPath) -> None:
    lowered_parts = [part.lower() for part in posix_uri.parts]
    if _FORBIDDEN_DIRECTORY_NAME in lowered_parts or posix_uri.name.lower() == _FORBIDDEN_FILENAME:
        raise ArtifactIngestionError(f"Evidence manifest uri '{uri}' cannot reference test-only ground truth.")


def _resolve_and_validate_uri(artifact_root: Path, uri: str) -> Path:
    resolved_root = artifact_root.resolve()
    resolved_path = (artifact_root / uri).resolve()
    if not resolved_path.is_relative_to(resolved_root):
        raise ArtifactIngestionError(f"Evidence manifest uri '{uri}' escapes the artifact root.")
    return resolved_path


def _read_verified_bytes(resolved_path: Path, expected_sha256: str) -> bytes:
    if not resolved_path.is_file():
        raise ArtifactIngestionError(f"Artifact file not found at {resolved_path}.")
    data = resolved_path.read_bytes()
    actual_sha256 = hashlib.sha256(data).hexdigest()
    if actual_sha256 != expected_sha256:
        raise ArtifactIngestionError(
            f"Checksum mismatch for {resolved_path}: expected {expected_sha256}, got {actual_sha256}."
        )
    return data


def _parse_artifact(source_type: str, data: bytes) -> str:
    try:
        return _PARSERS[source_type](data)
    except ArtifactIngestionError:
        raise
    except Exception as exc:
        raise ArtifactIngestionError(f"Failed to parse artifact of type '{source_type}'.") from exc


def _normalize_text(text: str) -> str:
    normalized = text.replace("\r\n", "\n").replace("\r", "\n")
    lines = [line.rstrip() for line in normalized.split("\n")]
    return "\n".join(lines).strip()


def _parse_email(data: bytes) -> str:
    message = BytesParser(policy=policy.default).parsebytes(data)
    subject = (message["Subject"] or "").strip()
    body = (message.get_content() or "").strip()
    if not subject and not body:
        raise ArtifactIngestionError("Email artifact has no subject or body content.")
    return _normalize_text(f"Subject: {subject}\n\n{body}")


def _parse_calendar(data: bytes) -> str:
    calendar = Calendar.from_ical(data)
    events = [component for component in calendar.walk() if component.name == "VEVENT"]
    if len(events) != 1:
        raise ArtifactIngestionError("Expected exactly one VEVENT in the production calendar artifact.")
    event = events[0]

    summary = _ics_text(event.get("SUMMARY"))
    location = _ics_text(event.get("LOCATION"))
    description = _ics_text(event.get("DESCRIPTION"))
    if not (summary or location or description):
        raise ArtifactIngestionError("Calendar artifact has no meaningful SUMMARY, LOCATION, or DESCRIPTION.")

    lines = []
    if summary:
        lines.append(f"Summary: {summary}")
    if location:
        lines.append(f"Location: {location}")
    if description:
        lines.append(f"Description: {description}")
    return _normalize_text("\n".join(lines))


def _ics_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_spreadsheet(data: bytes) -> str:
    workbook = load_workbook(io.BytesIO(data), data_only=True)
    sheet = workbook.active
    lines = [
        " | ".join(str(cell) for cell in row if cell is not None)
        for row in sheet.iter_rows(values_only=True)
    ]
    lines = [line for line in lines if line]
    content = _normalize_text("\n".join(lines))
    if not content:
        raise ArtifactIngestionError("Spreadsheet artifact has no meaningful cell values.")
    return content


def _parse_pdf(data: bytes) -> str:
    reader = PdfReader(io.BytesIO(data))
    pages_text = [page.extract_text() or "" for page in reader.pages]
    content = _normalize_text("\n".join(pages_text))
    if not content:
        raise ArtifactIngestionError("PDF artifact has no extractable text content.")
    return content


def _parse_markdown(data: bytes) -> str:
    content = _normalize_text(data.decode("utf-8"))
    if not content:
        raise ArtifactIngestionError("Markdown artifact has no meaningful content.")
    return content


_PARSERS = {
    "email": _parse_email,
    "calendar": _parse_calendar,
    "spreadsheet": _parse_spreadsheet,
    "pdf": _parse_pdf,
    "markdown": _parse_markdown,
}
